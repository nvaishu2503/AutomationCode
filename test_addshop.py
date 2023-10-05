import openpyxl
import pytest
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select

@pytest.mark.usefixtures("setup")

def test_vshopadd(setup):
    path1 = r'C:\Users\User\Desktop\vaishu test cases\vshop(shop details).xlsx'
    wb1 = openpyxl.load_workbook(path1)
    s = wb1.active
    row_ct = s.max_row
    col_ct = s.max_column

    for i in range(2, row_ct + 1):
        if s.cell(row=i, column=1).value == "Trends":
            r = []
            setup.implicitly_wait(10)
            setup.maximize_window()
            for j in range(1, col_ct + 1):
                r.append(s.cell(row=i, column=j).value)
                if j == 11:
                    setup.find_element(By.ID, "Name").send_keys(r[0])
                    setup.find_element(By.ID, "Email").send_keys(r[1])
                    setup.find_element(By.ID, "User Name").send_keys(r[2])
                    setup.find_element(By.ID, "Password").send_keys(r[3])
                    setup.find_element(By.ID, "Shop Name").send_keys(r[4])
                    dropdown = Select(setup.find_element(By.ID, "Category"))
                    dropdown.select_by_index(2)
                    setup.find_element(By.ID, "Description").send_keys("")
                    element = setup.find_element(By.ID, "Address Line 1")
                    setup.execute_script("return arguments[0].scrollIntoView(true);", element)
                    setup.find_element(By.ID, "Address Line 1").send_keys(r[5])
                    setup.find_element(By.ID, "Address Line 2").send_keys(r[6])
                    setup.find_element(By.ID, "City").send_keys(r[7])
                    setup.find_element(By.ID, "State").send_keys(r[8])
                    setup.find_element(By.ID, "Pincode").send_keys(r[9])
                    setup.find_element(By.ID, "Description").send_keys(r[10])
                    #close = setup.find_element(By.XPATH, "//body/div/div/div/div/div/div[2]")
                    #setup.execute_script("arguments[0].click();", close)
                    #save = setup.find_element(By.XPATH, "//button[normalize-space()='Save']")
                    #setup.execute_script("arguments[0].click();", save)
                    print(j)
                print(r)
