import pytest
from selenium import webdriver
import time
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
import openpyxl


options = webdriver.ChromeOptions()
options.add_experimental_option('detach', True)
service_obj = Service("E:\chromedriver\chromedriver.exe")
driver = webdriver.Chrome(options=options,service=service_obj)
driver.implicitly_wait(10)
driver.get("http://localhost:3000/admin")
driver.maximize_window()
driver.find_element(By.ID,"User Name").send_keys("newyorkerFin")
driver.find_element(By.ID,"Password").send_keys("Product$123")
driver.find_element(By.CSS_SELECTOR,"button[type='submit']").click()
driver.find_element(By.XPATH,"//a[@href='/admin/categories']").click()
driver.find_element(By.XPATH,"//button[normalize-space()='Add Category']").click()


path1 = r'C:\Users\User\Desktop\vaishu test cases\vshopcatergories.xlsx'
wb1 = openpyxl.load_workbook(path1)
s = wb1.active
#row_ct=s.max_row
#col_ct=s.max_column
c = s.cell(row=1,column=4)
print(c.value)
driver.find_element(By.XPATH,"//body/div/div/div/div/div/div/div/div/div/div[2]/div[1]").click()
driver.find_element(By.ID, "EnglishName").send_keys(c.value)
driver.find_element(By.XPATH,"//button[@type='submit']").click()


""""
for i in range(1, row_ct+1):
    r=[]
    for j in range(3, col_ct+1):
        r.append(s.cell(row=i,column=j).value)
        if i == 9:
            driver.find_element(By.XPATH,"//body/div/div/div/div/div/div/div/div/div/div[2]/div[1]").click()
            driver.find_element(By.ID, "English name").send_keys(r[0])
            #driver.find_element(By.XPATH,"//body//div//div//div//div//div[1]//div[1]//div[1]//div[2]//div[1]//div[1]").click()
            driver.find_element(By.XPATH,"//button[@type='submit']").click()
            #driver.find_element(By.XPATH, "//button[normalize-space()='Add Category']").click()
        print(r)
"""

#//body/div/div/div/div/div/div/div/div/div/div[2]/div[1]
#//button[@type='submit']