from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
import openpyxl

options = webdriver.ChromeOptions()
options.add_experimental_option('detach', True)
service_obj = Service("E:\chromedriver\chromedriver.exe")
driver = webdriver.Chrome(options=options, service=service_obj)
driver.implicitly_wait(10)
driver.get("http://localhost:3000/super/shops")
driver.find_element(By.ID,"User Name").send_keys("admin")
driver.find_element(By.ID,"Password").send_keys("admin@123")
driver.find_element(By.CSS_SELECTOR,"button[type='submit']").click()
driver.find_element(By.XPATH,"//button[normalize-space()='Add Shop']").click()

path1 = r'C:\Users\User\Desktop\vaishu test cases\vshop(shop details).xlsx'
wb1 = openpyxl.load_workbook(path1)
s = wb1.active

row_ct = s.max_row
col_ct = s.max_column

for i in range(2, row_ct+1):
    #print('\n')
    if s.cell(row=i,column=1).value == "NewyorkerFin":
        r = []
        driver.implicitly_wait(20)
        driver.maximize_window()
        #driver.find_element(By.XPATH, "//button[normalize-space()='Add Shop']").click()
        for j in range(1, col_ct+1):
            #print(s.cell(row=i, column=j).value, end='')
            r.append(s.cell(row=i, column=j).value)
            if j == 11:
                driver.find_element(By.ID, "Name").send_keys(r[0])
                driver.find_element(By.ID, "Email").send_keys(r[1])
                driver.find_element(By.ID, "User Name").send_keys(r[2])
                driver.find_element(By.ID, "Password").send_keys(r[3])
                driver.find_element(By.ID, "Shop Name").send_keys(r[4])
                dropdown = Select(driver.find_element(By.ID, "Category"))
                dropdown.select_by_index(2)
                driver.find_element(By.ID, "Description").send_keys("")
                element = driver.find_element(By.ID, "Address Line 1")
                driver.execute_script("return arguments[0].scrollIntoView(true);", element)
                driver.find_element(By.ID, "Address Line 1").send_keys(r[5])
                driver.find_element(By.ID, "Address Line 2").send_keys(r[6])
                driver.find_element(By.ID, "City").send_keys(r[7])
                driver.find_element(By.ID, "State").send_keys(r[8])
                driver.find_element(By.ID, "Pincode").send_keys(r[9])
                driver.find_element(By.ID, "Description").send_keys(r[10])
                #close = driver.find_element(By.XPATH, "//body/div/div/div/div/div/div[2]")
                #driver.execute_script("arguments[0].click();", close)
                save = driver.find_element(By.XPATH, "//button[normalize-space()='Save']")
                driver.execute_script("arguments[0].click();", save)
                print(j)
            print(r)