import openpyxl
from selenium import webdriver
import time

from selenium.webdriver import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select

options = webdriver.ChromeOptions()
options.add_experimental_option('detach', True)
service_obj = Service("E:\chromedriver\chromedriver.exe")
driver = webdriver.Chrome(options=options,service=service_obj)
driver.implicitly_wait(10)
driver.get("http://localhost:3000/admin")
driver.maximize_window()
driver.find_element(By.ID,"User Name").send_keys("newyorkerfin")
driver.find_element(By.ID,"Password").send_keys("Product$123")
driver.find_element(By.CSS_SELECTOR,"button[type='submit']").click()
driver.find_element(By.XPATH,"//a[@href='/admin/products']//div//span").click()
driver.find_element(By.XPATH,"//body//div//div//div//div//div//div//div//button").click()
driver.find_element(By.XPATH,"//body//div//div//div//div//div[1]//div[1]//div[1]//div[2]//div[2]//div[1]").click()

path1 = r'C:\Users\User\Desktop\vaishu test cases\NewyorkerFin.xlsx'
wb1 = openpyxl.load_workbook(path1)
s = wb1.active
row_ct = s.max_row
col_ct = s.max_column

for i in range(2, row_ct + 1):
    if s.cell(row=i, column=1).value == "Leggings":
        r = []
        driver.implicitly_wait(20)
        driver.maximize_window()
        for j in range(1, col_ct + 1):
            r.append(s.cell(row=i, column=j).value)
            if j == 5:
                driver.find_element(By.ID,"EnglishName").send_keys(r[0])
                driver.find_element(By.CSS_SELECTOR,"textarea[placeholder='EnglishDescription']").send_keys(r[1])
                driver.find_element(By.ID,"Image").send_keys(Keys.CONTROL + "a")
                driver.find_element(By.ID,"Image").send_keys(Keys.DELETE)
                driver.find_element(By.ID,"Image").send_keys(r[2])
                #driver.find_element(By.XPATH,"//body/div/div/div/div/div/div/div/div/div/div[4]/div[1]")
                category = driver.find_element(By.XPATH,"//div[5]//div[1]//div[1]//div[2]//div[2]//div[1]")
                driver.execute_script("return arguments[0].scrollIntoView(true);", category)
                driver.execute_script("arguments[0].click();", category)
                #driver.find_elements(By.CLASS_NAME,"hstack gap-3")
                price=driver.find_element(By.XPATH,"//body/div/div/div/div/div/div/div/div/div/div[2]/div[1]/div[1]")
                driver.execute_script("return arguments[0].scrollIntoView(true);", price)
                driver.execute_script("arguments[0].click();", price)
                driver.find_element(By.CSS_SELECTOR,"input[placeholder='Prices']").send_keys(r[3])
                driver.find_element(By.CSS_SELECTOR,"input[placeholder='Offer price']").send_keys(r[4])
                driver.find_element(By.CSS_SELECTOR,"button[type='submit']").click()



#div[class*='selectable-tile_clickable']-to click the english tab
#women catergory-//div[5]//div[1]//div[1]//div[2]//div[2]//div[1]
#Men catergory-//div[5]//div[1]//div[1]//div[2]//div[1]//div[1]