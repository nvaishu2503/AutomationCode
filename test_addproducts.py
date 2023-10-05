import openpyxl
import pytest
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By


@pytest.mark.usefixtures("addproducts")

def test_addingproducts(addproducts):
    path1 = r'C:\Users\User\Desktop\vaishu test cases\NewyorkerFin.xlsx'
    wb1 = openpyxl.load_workbook(path1)
    s = wb1.active
    row_ct = s.max_row
    col_ct = s.max_column

    for i in range(2, row_ct + 1):
        if s.cell(row=i, column=1).value == "Leggings":
            r = []
            addproducts.implicitly_wait(20)
            addproducts.maximize_window()
            for j in range(1, col_ct + 1):
                r.append(s.cell(row=i, column=j).value)
                if j == 5:
                    addproducts.find_element(By.ID,"EnglishName").send_keys(r[0])
                    addproducts.find_element(By.CSS_SELECTOR,"textarea[placeholder='EnglishDescription']").send_keys(r[1])
                    addproducts.find_element(By.ID,"Image").send_keys(Keys.CONTROL + "a")
                    addproducts.find_element(By.ID,"Image").send_keys(Keys.DELETE)
                    addproducts.find_element(By.ID,"Image").send_keys(r[2])
                    #driver.find_element(By.XPATH,"//body/div/div/div/div/div/div/div/div/div/div[4]/div[1]")
                    category = addproducts.find_element(By.XPATH,"//div[5]//div[1]//div[1]//div[2]//div[2]//div[1]")
                    addproducts.execute_script("return arguments[0].scrollIntoView(true);", category)
                    addproducts.execute_script("arguments[0].click();", category)
                    #driver.find_elements(By.CLASS_NAME,"hstack gap-3")
                    price=addproducts.find_element(By.XPATH,"//body/div/div/div/div/div/div/div/div/div/div[2]/div[1]/div[1]")
                    addproducts.execute_script("return arguments[0].scrollIntoView(true);", price)
                    addproducts.execute_script("arguments[0].click();", price)
                    addproducts.find_element(By.CSS_SELECTOR,"input[placeholder='Prices']").send_keys(r[3])
                    addproducts.find_element(By.CSS_SELECTOR,"input[placeholder='Offer price']").send_keys(r[4])
                    addproducts.find_element(By.CSS_SELECTOR,"button[type='submit']").click()

